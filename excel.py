#Script to extract test execution progress from qmetry.
#Always download report in xlsx format.
#
import openpyxl as xl
import pandas as pd
from vincent.colors import brews
wb=xl.load_workbook('TestSuites (6).xlsx',data_only=True)
sheet = wb['Sheet0']
cell=sheet['A1']
sum_of_test_cases=[]
dictList= []
for row in range(2,sheet.max_row+1):
    cells=sheet.cell(row,12)
    sum_of_test_cases.append(cells.value)
    blocked= sheet.cell(row,4)
    omit = sheet.cell(row,5)
    failed = sheet.cell(row,6)
    not_run=sheet.cell(row,7)
    passed = sheet.cell(row,8)
    test_owner =sheet.cell(row,16)
    keys = ['Blocked', 'Omit', 'Failed', 'Not Run', 'Passed', 'Test Suite Owner']
    fields_values_list = [int(blocked.value),int(omit.value),int(failed.value),int(not_run.value),int(passed.value),str(test_owner.value)]
    #write value

    dictList.append(dict(zip(keys,fields_values_list)))
    #print values_dict
Round = sum([int(cells) for cells in sum_of_test_cases])
print("Total Test Case: " ,Round )
index = [dictList[i]['Test Suite Owner'] for i in range(len(dictList))] #x-axis
#extract all user data except test-owner name from list of dictionary.
print 'Test Suite Owners: ',index

for d in dictList:
    del d['Test Suite Owner']

data = dictList #refined data

# Create a Pandas dataframe from the data.
df = pd.DataFrame(data, index=index)

# Create a Pandas Excel writer using XlsxWriter as the engine.
excel_file = 'Pre-ESV-progress.xlsx'
sheet_name = 'Sheet1'

writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
df.to_excel(writer, sheet_name=sheet_name)

# Access the XlsxWriter workbook and worksheet objects from the dataframe.
workbook = writer.book
worksheet = writer.sheets[sheet_name]

# Create a chart object.
chart = workbook.add_chart({'type': 'column','subtype': 'stacked'})

# Configure the series of the chart from the dataframe data.

for col_num in range(1, len(data) - 5):
        chart.add_series({
        'name':       ['Sheet1', 0, col_num],
        'categories': ['Sheet1', 1, 0, 14, 0],
        'values':     ['Sheet1', 1, col_num, 14, col_num],
        'fill':       {'color': brews['Pastel1'][col_num - 1]},
        'gap':        20,
    })
chart.set_chartarea({
    'border': {'none': True},
    'fill':   {'color': 'white'}
})
chart.set_plotarea({
    'border': {'color': 'black', 'width': 1, 'dash_type': 'dash'},
    'fill':   {'color': '#ffffff'}
})
chart.set_size({'width': 720, 'height': 476})

# Configure the chart axes.
chart.set_x_axis({'name': 'Test Suite Owners'})
chart.set_y_axis({'name': 'Number of Test Cases', 'major_gridlines': {'visible': False}})

# Insert the chart into the worksheet.
worksheet.insert_chart('H2', chart)

# Close the Pandas Excel writer and output the Excel file.
writer.save()


